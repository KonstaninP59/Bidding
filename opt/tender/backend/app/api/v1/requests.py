from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import load_workbook
from io import BytesIO

from app.core.database import get_db
from app.api.v1.deps import require_permissions
from app.models.request import Request, RequestItem
from app.schemas.request import RequestCreate, RequestOut
from app.services.audit import log_action
from app.core.enums import RequestStatus

router = APIRouter(prefix="/requests", tags=["Requests"])


@router.post("/", response_model=RequestOut)
async def create_request(
    payload: RequestCreate,
    db: AsyncSession = Depends(get_db),
    user=Depends(require_permissions(["manage_requests"])),
):
    request = Request(
        number=f"REQ-{payload.category_id}-{payload.subject[:5]}",
        subject=payload.subject,
        description=payload.description,
        category_id=payload.category_id,
    )

    for item in payload.items:
        request.items.append(RequestItem(**item.model_dump()))

    db.add(request)
    await db.commit()
    await db.refresh(request)

    await log_action(db, user.id, "CREATE", {"request_id": request.id})
    await db.commit()

    return request


@router.post("/{request_id}/import-items")
async def import_items(
    request_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user=Depends(require_permissions(["manage_requests"])),
):
    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    request = await db.get(Request, request_id)
    if not request:
        raise HTTPException(404)

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, description, qty, unit, *_ = row

        if not name or not qty or not unit:
            continue

        request.items.append(
            RequestItem(
                name=name,
                description=description,
                quantity=float(qty),
                unit=unit,
            )
        )

    await db.commit()
    return {"status": "imported"}
