import os
import io
import re
import json
import hmac
import time
import base64
import hashlib
import secrets
import smtplib
import argparse
from datetime import datetime, date, timedelta, timezone
from email.message import EmailMessage
from typing import Optional, List, Dict, Tuple

from dotenv import load_dotenv

# Matplotlib headless backend
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from fastapi import (
    FastAPI, Request, Form, Depends, HTTPException, status,
    UploadFile, File, BackgroundTasks
)
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse, PlainTextResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles

from starlette.middleware.sessions import SessionMiddleware

from passlib.context import CryptContext

from sqlalchemy import (
    create_engine, String, Integer, DateTime, Boolean, ForeignKey, Text, Numeric, LargeBinary,
    UniqueConstraint, Index, select, func, text
)
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, relationship, Session

from openpyxl import load_workbook, Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet


# -----------------------------
# Config
# -----------------------------
load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL", "")
SECRET_KEY = os.getenv("SECRET_KEY", "dev-secret-change-me")
BASE_URL = os.getenv("BASE_URL", "http://127.0.0.1:8000").rstrip("/")

SMTP_HOST = os.getenv("SMTP_HOST", "").strip()
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "").strip()
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "").strip()
SMTP_FROM = os.getenv("SMTP_FROM", "procurement@example.com").strip()
SMTP_TLS = os.getenv("SMTP_TLS", "true").lower() in ("1", "true", "yes", "on")

UPLOAD_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "uploads"))
TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates")

MAX_UPLOAD_MB = 50  # просто и понятно; можно вынести в настройки
MAX_UPLOAD_BYTES = MAX_UPLOAD_MB * 1024 * 1024

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


# -----------------------------
# DB
# -----------------------------
class Base(DeclarativeBase):
    pass


class Role:
    ADMIN = "ADMIN"
    BUYER = "BUYER"
    VIEWER = "VIEWER"


class RequestStatus:
    DRAFT = "DRAFT"
    ROUND_1_OPEN = "ROUND_1_OPEN"
    ROUND_1_CLOSED = "ROUND_1_CLOSED"
    ROUND_N_OPEN = "ROUND_N_OPEN"
    ROUND_N_CLOSED = "ROUND_N_CLOSED"
    DECISION = "DECISION"
    COMPLETED = "COMPLETED"
    CANCELLED = "CANCELLED"


class RoundType:
    INITIAL = "INITIAL"
    NEGOTIATION = "NEGOTIATION"


class SupplierStatus:
    ACTIVE = "ACTIVE"
    PAUSE = "PAUSE"
    BLOCKED = "BLOCKED"


class InvitationStatus:
    CREATED = "CREATED"
    SENT = "SENT"
    OPENED = "OPENED"
    RESPONDED = "RESPONDED"
    EXPIRED = "EXPIRED"


class User(Base):
    __tablename__ = "users"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    email: Mapped[str] = mapped_column(String(320), unique=True, index=True)
    name: Mapped[str] = mapped_column(String(200), default="")
    password_hash: Mapped[str] = mapped_column(String(255))
    role: Mapped[str] = mapped_column(String(20), default=Role.VIEWER)
    is_active: Mapped[bool] = mapped_column(Boolean, default=True)

    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class CategorySupplier(Base):
    __tablename__ = "category_suppliers"
    category_id: Mapped[int] = mapped_column(ForeignKey("categories.id"), primary_key=True)
    supplier_id: Mapped[int] = mapped_column(ForeignKey("suppliers.id"), primary_key=True)


class RoundSupplier(Base):
    __tablename__ = "round_suppliers"
    round_id: Mapped[int] = mapped_column(ForeignKey("rounds.id"), primary_key=True)
    supplier_id: Mapped[int] = mapped_column(ForeignKey("suppliers.id"), primary_key=True)


class Category(Base):
    __tablename__ = "categories"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    name: Mapped[str] = mapped_column(String(200), unique=True, index=True)
    description: Mapped[str] = mapped_column(Text, default="")
    # Простая настройка обязательности полей формы КП (можно расширять)
    require_payment_terms: Mapped[bool] = mapped_column(Boolean, default=False)
    allow_not_supply: Mapped[bool] = mapped_column(Boolean, default=True)
    forbid_price_increase: Mapped[bool] = mapped_column(Boolean, default=True)

    suppliers: Mapped[List["Supplier"]] = relationship(
        secondary="category_suppliers", back_populates="categories"
    )


class Supplier(Base):
    __tablename__ = "suppliers"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    name: Mapped[str] = mapped_column(String(250), unique=True, index=True)
    emails: Mapped[str] = mapped_column(Text, default="")  # comma-separated
    contact_name: Mapped[str] = mapped_column(String(200), default="")
    phone: Mapped[str] = mapped_column(String(50), default="")
    status: Mapped[str] = mapped_column(String(20), default=SupplierStatus.ACTIVE)

    categories: Mapped[List[Category]] = relationship(
        secondary="category_suppliers", back_populates="suppliers"
    )


class PurchaseRequest(Base):
    __tablename__ = "requests"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    number: Mapped[str] = mapped_column(String(50), unique=True, index=True)
    category_id: Mapped[int] = mapped_column(ForeignKey("categories.id"))
    subject: Mapped[str] = mapped_column(String(300))
    description: Mapped[str] = mapped_column(Text, default="")
    currency: Mapped[str] = mapped_column(String(10), default="RUB")
    vat_mode: Mapped[str] = mapped_column(String(50), default="NDS_INCLUDED")  # простая строка
    payment_terms: Mapped[str] = mapped_column(Text, default="")
    delivery_terms: Mapped[str] = mapped_column(Text, default="")

    status: Mapped[str] = mapped_column(String(30), default=RequestStatus.DRAFT)

    created_by_user_id: Mapped[int] = mapped_column(ForeignKey("users.id"))
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))

    category: Mapped[Category] = relationship()
    items: Mapped[List["RequestItem"]] = relationship(back_populates="request", cascade="all, delete-orphan")
    rounds: Mapped[List["Round"]] = relationship(back_populates="request", cascade="all, delete-orphan")


class RequestItem(Base):
    __tablename__ = "request_items"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    request_id: Mapped[int] = mapped_column(ForeignKey("requests.id"), index=True)
    pos_no: Mapped[int] = mapped_column(Integer, default=1)
    name: Mapped[str] = mapped_column(String(300))
    description: Mapped[str] = mapped_column(Text, default="")
    qty: Mapped[float] = mapped_column(Numeric(14, 3))
    uom: Mapped[str] = mapped_column(String(50), default="шт")
    required_date: Mapped[Optional[date]] = mapped_column(DateTime(timezone=False), nullable=True)

    request: Mapped[PurchaseRequest] = relationship(back_populates="items")


class Round(Base):
    __tablename__ = "rounds"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    request_id: Mapped[int] = mapped_column(ForeignKey("requests.id"), index=True)

    number: Mapped[int] = mapped_column(Integer)  # 1,2,...
    type: Mapped[str] = mapped_column(String(20), default=RoundType.INITIAL)
    deadline_at: Mapped[datetime] = mapped_column(DateTime(timezone=True))
    comment_to_suppliers: Mapped[str] = mapped_column(Text, default="")
    is_closed: Mapped[bool] = mapped_column(Boolean, default=False)
    closed_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True), nullable=True)

    created_by_user_id: Mapped[int] = mapped_column(ForeignKey("users.id"))
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))

    request: Mapped[PurchaseRequest] = relationship(back_populates="rounds")
    invitations: Mapped[List["Invitation"]] = relationship(back_populates="round", cascade="all, delete-orphan")
    offers: Mapped[List["Offer"]] = relationship(back_populates="round", cascade="all, delete-orphan")

    __table_args__ = (UniqueConstraint("request_id", "number", name="uq_round_request_number"),)


class Invitation(Base):
    __tablename__ = "invitations"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    round_id: Mapped[int] = mapped_column(ForeignKey("rounds.id"), index=True)
    supplier_id: Mapped[int] = mapped_column(ForeignKey("suppliers.id"), index=True)

    token_hash: Mapped[str] = mapped_column(String(64), index=True)  # sha256 hex
    status: Mapped[str] = mapped_column(String(20), default=InvitationStatus.CREATED)

    sent_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True), nullable=True)
    opened_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True), nullable=True)
    responded_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True), nullable=True)
    reminded_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True), nullable=True)

    round: Mapped[Round] = relationship(back_populates="invitations")
    supplier: Mapped[Supplier] = relationship()

    __table_args__ = (UniqueConstraint("round_id", "supplier_id", name="uq_inv_round_supplier"),)


class Offer(Base):
    __tablename__ = "offers"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    round_id: Mapped[int] = mapped_column(ForeignKey("rounds.id"), index=True)
    supplier_id: Mapped[int] = mapped_column(ForeignKey("suppliers.id"), index=True)

    payment_terms: Mapped[str] = mapped_column(Text, default="")
    valid_until: Mapped[str] = mapped_column(String(100), default="")
    comment: Mapped[str] = mapped_column(Text, default="")

    total_amount: Mapped[Optional[float]] = mapped_column(Numeric(18, 2), nullable=True)

    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))

    round: Mapped[Round] = relationship(back_populates="offers")
    supplier: Mapped[Supplier] = relationship()
    items: Mapped[List["OfferItem"]] = relationship(back_populates="offer", cascade="all, delete-orphan")

    __table_args__ = (UniqueConstraint("round_id", "supplier_id", name="uq_offer_round_supplier"),)


class OfferItem(Base):
    __tablename__ = "offer_items"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    offer_id: Mapped[int] = mapped_column(ForeignKey("offers.id"), index=True)
    request_item_id: Mapped[int] = mapped_column(ForeignKey("request_items.id"), index=True)

    unit_price: Mapped[float] = mapped_column(Numeric(18, 2))
    delivery_days: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)
    comment: Mapped[str] = mapped_column(Text, default="")
    not_supply: Mapped[bool] = mapped_column(Boolean, default=False)

    line_amount: Mapped[Optional[float]] = mapped_column(Numeric(18, 2), nullable=True)

    offer: Mapped[Offer] = relationship(back_populates="items")


class Attachment(Base):
    __tablename__ = "attachments"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    kind: Mapped[str] = mapped_column(String(30))  # REQUEST, OFFER
    request_id: Mapped[Optional[int]] = mapped_column(ForeignKey("requests.id"), nullable=True, index=True)
    offer_id: Mapped[Optional[int]] = mapped_column(ForeignKey("offers.id"), nullable=True, index=True)

    original_name: Mapped[str] = mapped_column(String(400))
    stored_path: Mapped[str] = mapped_column(String(500))
    content_type: Mapped[str] = mapped_column(String(200), default="")
    size_bytes: Mapped[int] = mapped_column(Integer)

    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class Report(Base):
    __tablename__ = "reports"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    request_id: Mapped[int] = mapped_column(ForeignKey("requests.id"), index=True)
    round_id: Mapped[int] = mapped_column(ForeignKey("rounds.id"), index=True)

    kind: Mapped[str] = mapped_column(String(10))  # PDF or XLSX
    file_bytes: Mapped[bytes] = mapped_column(LargeBinary)
    filename: Mapped[str] = mapped_column(String(200))
    snapshot_json: Mapped[str] = mapped_column(Text, default="{}")

    created_by_user_id: Mapped[int] = mapped_column(ForeignKey("users.id"))
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class AuditLog(Base):
    __tablename__ = "audit_log"
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), index=True)
    actor_user_id: Mapped[Optional[int]] = mapped_column(ForeignKey("users.id"), nullable=True, index=True)
    actor_supplier_id: Mapped[Optional[int]] = mapped_column(ForeignKey("suppliers.id"), nullable=True, index=True)
    ip: Mapped[str] = mapped_column(String(60), default="")
    action: Mapped[str] = mapped_column(String(80))
    entity: Mapped[str] = mapped_column(String(80), default="")
    entity_id: Mapped[Optional[int]] = mapped_column(Integer, nullable=True)
    details: Mapped[str] = mapped_column(Text, default="")

Index("ix_requests_status_created", PurchaseRequest.status, PurchaseRequest.created_at)
Index("ix_requests_category_created", PurchaseRequest.category_id, PurchaseRequest.created_at)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)


def db_session():
    with Session(engine) as s:
        yield s


# -----------------------------
# App + templates
# -----------------------------
app = FastAPI(title="Tender MVP", version="1.0")
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY, same_site="lax")

templates = Jinja2Templates(directory=TEMPLATES_DIR)

# uploads as static (для скачивания вложений)
os.makedirs(UPLOAD_DIR, exist_ok=True)
# app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")


def _attachment_file_path(att: Attachment) -> str:
    # поддержка старых записей вида "/uploads/xxx" и новых "xxx"
    sp = att.stored_path or ""
    sp = sp.replace("\\", "/")
    if sp.startswith("/uploads/"):
        sp = sp.split("/uploads/", 1)[1]
    return os.path.join(UPLOAD_DIR, sp)


@app.get("/attachments/{attachment_id}/download")
def attachment_download(
    request: Request,
    attachment_id: int,
    token: str = "",
    db: Session = Depends(db_session),
):
    att = db.get(Attachment, attachment_id)
    if not att:
        raise HTTPException(404, "Attachment not found")

    # 1) если залогинен — разрешаем
    user_id = request.session.get("user_id")
    if user_id:
        pass
    else:
        # 2) иначе — только по token (поставщик)
        if not token:
            raise HTTPException(401, "Not authorized")
        inv = find_invitation_by_token(db, token)
        rnd = db.get(Round, inv.round_id)
        req = db.get(PurchaseRequest, rnd.request_id)

        # supplier can download:
        # - request attachments of this request
        # - own offer attachments for this round
        if att.kind == "REQUEST":
            if att.request_id != req.id:
                raise HTTPException(403, "Forbidden")
        elif att.kind == "OFFER":
            off = db.get(Offer, att.offer_id) if att.offer_id else None
            if not off or off.supplier_id != inv.supplier_id or off.round_id != rnd.id:
                raise HTTPException(403, "Forbidden")
        else:
            raise HTTPException(403, "Forbidden")

    file_path = _attachment_file_path(att)
    if not os.path.exists(file_path):
        raise HTTPException(404, "File missing on disk")

    headers = {"Content-Disposition": f'attachment; filename="{safe_filename(att.original_name)}"'}
    return StreamingResponse(open(file_path, "rb"), media_type=att.content_type or "application/octet-stream", headers=headers)


# -----------------------------
# Helpers: auth/RBAC
# -----------------------------
def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def require_login(request: Request, db: Session = Depends(db_session)) -> User:
    user_id = request.session.get("user_id")
    if not user_id:
        raise HTTPException(status_code=401, detail="Not logged in")
    user = db.get(User, int(user_id))
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="User inactive")
    return user


def require_roles(*roles: str):
    def dep(user: User = Depends(require_login)) -> User:
        if user.role not in roles:
            raise HTTPException(status_code=403, detail="Forbidden")
        return user
    return dep


def hash_password(p: str) -> str:
    return pwd_context.hash(p)


def verify_password(p: str, hashed: str) -> bool:
    return pwd_context.verify(p, hashed)


def audit(db: Session, request: Request, action: str, entity: str = "", entity_id: Optional[int] = None, details: str = "", user: Optional[User] = None, supplier: Optional[Supplier] = None):
    ip = request.client.host if request.client else ""
    db.add(AuditLog(
        actor_user_id=user.id if user else None,
        actor_supplier_id=supplier.id if supplier else None,
        ip=ip,
        action=action,
        entity=entity,
        entity_id=entity_id,
        details=details[:5000]
    ))
    db.commit()


# -----------------------------
# Helpers: invitations/tokens
# -----------------------------
def token_hash(token: str) -> str:
    # Храним только sha256(token + SECRET_KEY) (не просто token)
    h = hashlib.sha256()
    h.update((token + SECRET_KEY).encode("utf-8"))
    return h.hexdigest()


def make_public_link(token: str) -> str:
    return f"{BASE_URL}/public/offer/{token}"


# -----------------------------
# Helpers: rate limiting (простое, in-memory)
# -----------------------------
RATE_BUCKET: Dict[str, List[float]] = {}  # ip -> timestamps
RATE_LIMIT = 60  # req/min per ip for public form


def rate_limit_public(request: Request):
    ip = request.client.host if request.client else "unknown"
    t = time.time()
    arr = RATE_BUCKET.get(ip, [])
    arr = [x for x in arr if (t - x) < 60]
    if len(arr) >= RATE_LIMIT:
        raise HTTPException(status_code=429, detail="Too many requests")
    arr.append(t)
    RATE_BUCKET[ip] = arr


# -----------------------------
# Helpers: calculations
# -----------------------------
def compute_offer_totals(db: Session, offer: Offer) -> float:
    total = 0.0
    for oi in offer.items:
        ri = db.get(RequestItem, oi.request_item_id)
        qty = float(ri.qty)
        if oi.not_supply:
            oi.line_amount = 0
            continue
        line = qty * float(oi.unit_price)
        oi.line_amount = round(line, 2)
        total += line
    offer.total_amount = round(total, 2)
    return float(offer.total_amount)


def get_round_offer(db: Session, round_id: int, supplier_id: int) -> Optional[Offer]:
    return db.scalar(select(Offer).where(Offer.round_id == round_id, Offer.supplier_id == supplier_id))


def round_is_open(r: Round) -> bool:
    if r.is_closed:
        return False
    return now_utc() <= r.deadline_at


def close_round_if_deadline_passed(db: Session, r: Round):
    if not r.is_closed and now_utc() > r.deadline_at:
        r.is_closed = True
        r.closed_at = now_utc()
        db.commit()


# -----------------------------
# Email sending
# -----------------------------
def send_email_smtp(to_email: str, subject: str, text: str):
    msg = EmailMessage()
    msg["From"] = SMTP_FROM
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(text)

    if not SMTP_HOST:
        # fallback: dev mode
        print("\n--- EMAIL (dev) ---")
        print("TO:", to_email)
        print("SUBJECT:", subject)
        print(text)
        print("--- END ---\n")
        return

    if SMTP_TLS:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls()
            if SMTP_USER:
                s.login(SMTP_USER, SMTP_PASSWORD)
            s.send_message(msg)
    else:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            if SMTP_USER:
                s.login(SMTP_USER, SMTP_PASSWORD)
            s.send_message(msg)


def build_invite_email(req: PurchaseRequest, rnd: Round, link: str) -> Tuple[str, str]:
    subject = f"Запрос КП №{req.number} / Раунд {rnd.number}"
    deadline_str = rnd.deadline_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    text = (
        "Здравствуйте!\n\n"
        f"Просим предоставить коммерческое предложение по заявке №{req.number} ({req.subject}).\n\n"
        f"Дедлайн: {deadline_str}\n"
        f"Ссылка для заполнения предложения: {link}\n\n"
        "Если ссылка не открывается — проверьте, что дедлайн не истёк.\n"
    )
    if rnd.comment_to_suppliers:
        text += f"\nКомментарий закупщика:\n{rnd.comment_to_suppliers}\n"
    return subject, text


def build_reminder_email(req: PurchaseRequest, rnd: Round, link: str) -> Tuple[str, str]:
    subject = f"Напоминание: КП №{req.number} до {rnd.deadline_at.strftime('%Y-%m-%d %H:%M')}"
    text = (
        "Здравствуйте!\n\n"
        f"Напоминаем о необходимости предоставить КП по заявке №{req.number}.\n"
        f"Раунд: {rnd.number}\n"
        f"Дедлайн: {rnd.deadline_at.strftime('%Y-%m-%d %H:%M %Z')}\n\n"
        f"Ссылка: {link}\n"
    )
    return subject, text


# -----------------------------
# Startup init
# -----------------------------
@app.on_event("startup")
def startup():
    Base.metadata.create_all(engine)

    with engine.connect() as conn:
        conn.execute(text("ALTER TABLE invitations ADD COLUMN IF NOT EXISTS reminded_at TIMESTAMPTZ NULL;"))
        conn.commit()


# -----------------------------
# Web: auth
# -----------------------------
@app.get("/", response_class=HTMLResponse)
def root(request: Request, user: Optional[User] = Depends(lambda: None)):
    # Если залогинен — на дашборд, иначе — на логин
    if request.session.get("user_id"):
        return RedirectResponse("/dashboard", status_code=303)
    return RedirectResponse("/login", status_code=303)


@app.get("/admin/users", response_class=HTMLResponse)
def admin_users_page(
    request: Request,
    user: User = Depends(require_roles(Role.ADMIN)),
    db: Session = Depends(db_session),
):
    users = db.scalars(select(User).order_by(User.created_at.desc())).all()
    return templates.TemplateResponse("users.html", {"request": request, "user": user, "users": users, "Role": Role})


@app.post("/admin/users/new")
def admin_users_new(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    role: str = Form(Role.VIEWER),
    name: str = Form(""),
    user: User = Depends(require_roles(Role.ADMIN)),
    db: Session = Depends(db_session),
):
    email = email.strip().lower()
    role = role.strip().upper()

    if role not in (Role.ADMIN, Role.BUYER, Role.VIEWER):
        raise HTTPException(400, "Некорректная роль")

    if len(password.encode("utf-8")) > 72:
        raise HTTPException(400, "Пароль слишком длинный для bcrypt (<=72 байт)")

    if db.scalar(select(User).where(User.email == email)):
        raise HTTPException(400, "Пользователь с таким email уже существует")

    u = User(email=email, name=name.strip(), password_hash=hash_password(password), role=role, is_active=True)
    db.add(u)
    db.commit()
    audit(db, request, "USER_CREATE", "User", u.id, f"email={email}, role={role}", user=user)
    return RedirectResponse("/admin/users", status_code=303)


@app.post("/admin/users/{user_id}/toggle")
def admin_users_toggle(
    request: Request,
    user_id: int,
    user: User = Depends(require_roles(Role.ADMIN)),
    db: Session = Depends(db_session),
):
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(404, "User not found")
    if u.id == user.id:
        raise HTTPException(400, "Нельзя отключить самого себя")
    u.is_active = not u.is_active
    db.commit()
    audit(db, request, "USER_TOGGLE_ACTIVE", "User", u.id, f"is_active={u.is_active}", user=user)
    return RedirectResponse("/admin/users", status_code=303)


@app.post("/admin/users/{user_id}/role")
def admin_users_set_role(
    request: Request,
    user_id: int,
    role: str = Form(...),
    user: User = Depends(require_roles(Role.ADMIN)),
    db: Session = Depends(db_session),
):
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(404, "User not found")
    role = role.strip().upper()
    if role not in (Role.ADMIN, Role.BUYER, Role.VIEWER):
        raise HTTPException(400, "Некорректная роль")
    u.role = role
    db.commit()
    audit(db, request, "USER_SET_ROLE", "User", u.id, f"role={role}", user=user)
    return RedirectResponse("/admin/users", status_code=303)



@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login")
def login_post(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(db_session),
):
    user = db.scalar(select(User).where(User.email == email.lower().strip()))
    if not user or not user.is_active or not verify_password(password, user.password_hash):
        return templates.TemplateResponse("login.html", {"request": request, "error": "Неверный email или пароль"}, status_code=400)
    request.session["user_id"] = user.id
    return RedirectResponse("/dashboard", status_code=303)


@app.post("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


# -----------------------------
# Dashboard + search
# -----------------------------
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    q: str = "",
    status_filter: str = "",
    user: User = Depends(require_login),
    db: Session = Depends(db_session),
):
    stmt = select(PurchaseRequest).order_by(PurchaseRequest.created_at.desc())
    if q:
        # поиск по номеру/теме
        like = f"%{q.strip()}%"
        stmt = stmt.where((PurchaseRequest.number.ilike(like)) | (PurchaseRequest.subject.ilike(like)))
    if status_filter:
        stmt = stmt.where(PurchaseRequest.status == status_filter)

    reqs = db.scalars(stmt.limit(100)).all()

    # простые блоки "ближайшие дедлайны" и "DECISION"
    upcoming_rounds = db.scalars(
        select(Round).where(Round.is_closed == False).order_by(Round.deadline_at.asc()).limit(10)
    ).all()

    decision_reqs = db.scalars(
        select(PurchaseRequest).where(PurchaseRequest.status == RequestStatus.DECISION).order_by(PurchaseRequest.created_at.desc()).limit(10)
    ).all()

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": user,
        "reqs": reqs,
        "upcoming_rounds": upcoming_rounds,
        "decision_reqs": decision_reqs,
        "q": q,
        "status_filter": status_filter,
        "RequestStatus": RequestStatus
    })


# -----------------------------
# Directories: categories
# -----------------------------
@app.get("/categories", response_class=HTMLResponse)
def categories_page(request: Request, user: User = Depends(require_roles(Role.ADMIN, Role.BUYER, Role.VIEWER)), db: Session = Depends(db_session)):
    cats = db.scalars(select(Category).order_by(Category.name.asc())).all()
    sups = db.scalars(select(Supplier).order_by(Supplier.name.asc())).all()
    return templates.TemplateResponse("categories.html", {"request": request, "user": user, "cats": cats, "sups": sups})


@app.post("/categories/new")
def categories_new(
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    require_payment_terms: Optional[str] = Form(None),
    allow_not_supply: Optional[str] = Form(None),
    forbid_price_increase: Optional[str] = Form(None),
    supplier_ids: Optional[str] = Form(""),
    user: User = Depends(require_roles(Role.ADMIN)),
    db: Session = Depends(db_session),
):
    c = Category(
        name=name.strip(),
        description=description.strip(),
        require_payment_terms=bool(require_payment_terms),
        allow_not_supply=bool(allow_not_supply),
        forbid_price_increase=bool(forbid_price_increase),
    )
    db.add(c)
    db.commit()
    db.refresh(c)

    ids = [int(x) for x in re.findall(r"\d+", supplier_ids or "")]
    if ids:
        suppliers = db.scalars(select(Supplier).where(Supplier.id.in_(ids))).all()
        c.suppliers = suppliers
        db.commit()

    audit(db, request, "CATEGORY_CREATE", "Category", c.id, f"name={c.name}", user=user)
    return RedirectResponse("/categories", status_code=303)


# -----------------------------
# Directories: suppliers
# -----------------------------
@app.get("/suppliers", response_class=HTMLResponse)
def suppliers_page(request: Request, user: User = Depends(require_roles(Role.ADMIN, Role.BUYER, Role.VIEWER)), db: Session = Depends(db_session)):
    sups = db.scalars(select(Supplier).order_by(Supplier.name.asc())).all()
    cats = db.scalars(select(Category).order_by(Category.name.asc())).all()
    return templates.TemplateResponse("suppliers.html", {"request": request, "user": user, "sups": sups, "cats": cats, "SupplierStatus": SupplierStatus})


@app.post("/suppliers/new")
def suppliers_new(
    request: Request,
    name: str = Form(...),
    emails: str = Form(""),
    contact_name: str = Form(""),
    phone: str = Form(""),
    status_value: str = Form(SupplierStatus.ACTIVE),
    category_ids: str = Form(""),
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)),
    db: Session = Depends(db_session),
):
    s = Supplier(
        name=name.strip(),
        emails=",".join([e.strip() for e in emails.split(",") if e.strip()]),
        contact_name=contact_name.strip(),
        phone=phone.strip(),
        status=status_value
    )
    db.add(s)
    db.commit()
    db.refresh(s)

    ids = [int(x) for x in re.findall(r"\d+", category_ids or "")]
    if ids:
        cats = db.scalars(select(Category).where(Category.id.in_(ids))).all()
        s.categories = cats
        db.commit()

    audit(db, request, "SUPPLIER_CREATE", "Supplier", s.id, f"name={s.name}", user=user)
    return RedirectResponse("/suppliers", status_code=303)


# -----------------------------
# Requests: create/list
# -----------------------------
@app.get("/requests/new", response_class=HTMLResponse)
def request_new_page(request: Request, user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)), db: Session = Depends(db_session)):
    cats = db.scalars(select(Category).order_by(Category.name.asc())).all()
    return templates.TemplateResponse("request_new.html", {"request": request, "user": user, "cats": cats})


@app.post("/requests/new")
def request_new_post(
    request: Request,
    category_id: int = Form(...),
    subject: str = Form(...),
    description: str = Form(""),
    currency: str = Form("RUB"),
    vat_mode: str = Form("NDS_INCLUDED"),
    payment_terms: str = Form(""),
    delivery_terms: str = Form(""),
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)),
    db: Session = Depends(db_session),
):
    req = PurchaseRequest(
        number="TMP",
        category_id=category_id,
        subject=subject.strip(),
        description=description.strip(),
        currency=currency.strip(),
        vat_mode=vat_mode.strip(),
        payment_terms=payment_terms.strip(),
        delivery_terms=delivery_terms.strip(),
        created_by_user_id=user.id,
        status=RequestStatus.DRAFT
    )
    db.add(req)
    db.commit()
    db.refresh(req)

    # номер по id (максимально просто)
    req.number = f"REQ-{req.id:06d}"
    db.commit()

    audit(db, request, "REQUEST_CREATE", "Request", req.id, f"number={req.number}", user=user)
    return RedirectResponse(f"/requests/{req.id}", status_code=303)


@app.get("/requests/{request_id}", response_class=HTMLResponse)
def request_detail(
    request: Request,
    request_id: int,
    tab: str = "details",
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER, Role.VIEWER)),
    db: Session = Depends(db_session),
):
    req = db.get(PurchaseRequest, request_id)
    if not req:
        raise HTTPException(404, "Request not found")

    items = db.scalars(select(RequestItem).where(RequestItem.request_id == req.id).order_by(RequestItem.pos_no.asc())).all()
    rounds = db.scalars(select(Round).where(Round.request_id == req.id).order_by(Round.number.asc())).all()

    # Авто-закрытие при просрочке (простая реализация FR-13)
    for r in rounds:
        close_round_if_deadline_passed(db, r)

    rounds = db.scalars(select(Round).where(Round.request_id == req.id).order_by(Round.number.asc())).all()

    # предложения по выбранному раунду
    selected_round_id = None
    offers = []
    compare = None
    if rounds:
        selected_round_id = int(request.query_params.get("round_id", rounds[-1].id))
        offers = db.scalars(select(Offer).where(Offer.round_id == selected_round_id).order_by(Offer.total_amount.asc().nulls_last())).all()
        compare = build_compare_table(db, req, selected_round_id)

    attachments = db.scalars(select(Attachment).where(Attachment.request_id == req.id).order_by(Attachment.created_at.desc())).all()
    reports = db.scalars(select(Report).where(Report.request_id == req.id).order_by(Report.created_at.desc())).all()
    audit_logs = db.scalars(select(AuditLog).where(AuditLog.entity == "Request", AuditLog.entity_id == req.id).order_by(AuditLog.at.desc()).limit(200)).all()

    cat = db.get(Category, req.category_id)

    return templates.TemplateResponse("request_detail.html", {
        "request": request,
        "user": user,
        "req": req,
        "cat": cat,
        "items": items,
        "rounds": rounds,
        "selected_round_id": selected_round_id,
        "offers": offers,
        "compare": compare,
        "attachments": attachments,
        "reports": reports,
        "audit_logs": audit_logs,
        "tab": tab,
        "RequestStatus": RequestStatus,
        "RoundType": RoundType
    })


@app.post("/requests/{request_id}/items/add")
def request_item_add(
    request: Request,
    request_id: int,
    name: str = Form(...),
    description: str = Form(""),
    qty: float = Form(...),
    uom: str = Form("шт"),
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)),
    db: Session = Depends(db_session),
):
    req = db.get(PurchaseRequest, request_id)
    if not req:
        raise HTTPException(404, "Request not found")

    max_pos = db.scalar(select(func.max(RequestItem.pos_no)).where(RequestItem.request_id == request_id)) or 0
    ri = RequestItem(
        request_id=request_id,
        pos_no=int(max_pos) + 1,
        name=name.strip(),
        description=description.strip(),
        qty=qty,
        uom=uom.strip()
    )
    db.add(ri)
    db.commit()
    audit(db, request, "REQUEST_ITEM_ADD", "Request", req.id, f"item={ri.name}", user=user)
    return RedirectResponse(f"/requests/{request_id}?tab=spec", status_code=303)


@app.post("/requests/{request_id}/items/import_xlsx")
async def request_items_import_xlsx(
    request: Request,
    request_id: int,
    xlsx: UploadFile = File(...),
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)),
    db: Session = Depends(db_session),
):
    req = db.get(PurchaseRequest, request_id)
    if not req:
        raise HTTPException(404, "Request not found")

    data = await xlsx.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(400, f"Файл слишком большой (>{MAX_UPLOAD_MB}MB)")

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    # Ожидаем заголовки в первой строке
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[str(v).strip().lower()] = col

    def col_of(*names):
        for n in names:
            if n.lower() in headers:
                return headers[n.lower()]
        return None

    c_name = col_of("наименование", "name")
    c_desc = col_of("описание", "description")
    c_qty = col_of("количество", "qty")
    c_uom = col_of("ед.изм.", "ед", "uom")
    c_reqdate = col_of("требуемая дата", "required_date")

    if not (c_name and c_qty and c_uom):
        raise HTTPException(400, "Не найдены обязательные колонки: Наименование, Количество, Ед.изм.")

    max_pos = db.scalar(select(func.max(RequestItem.pos_no)).where(RequestItem.request_id == request_id)) or 0
    added = 0

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=c_name).value
        if not name:
            continue
        qty_val = ws.cell(row=row, column=c_qty).value
        uom_val = ws.cell(row=row, column=c_uom).value
        desc_val = ws.cell(row=row, column=c_desc).value if c_desc else ""
        req_date_val = ws.cell(row=row, column=c_reqdate).value if c_reqdate else None

        try:
            qty_f = float(qty_val)
        except Exception:
            raise HTTPException(400, f"Ошибка количества в строке {row}")

        ri = RequestItem(
            request_id=request_id,
            pos_no=int(max_pos) + 1 + added,
            name=str(name).strip(),
            description=str(desc_val).strip() if desc_val else "",
            qty=qty_f,
            uom=str(uom_val).strip() if uom_val else "шт",
        )
        db.add(ri)
        added += 1

    db.commit()
    audit(db, request, "REQUEST_ITEMS_IMPORT_XLSX", "Request", req.id, f"added={added}", user=user)
    return RedirectResponse(f"/requests/{request_id}?tab=spec", status_code=303)


# -----------------------------
# Attachments
# -----------------------------
def safe_filename(name: str) -> str:
    name = name.replace("\\", "/").split("/")[-1]
    name = re.sub(r"[^a-zA-Z0-9а-яА-Я._\- ()]+", "_", name)
    return name[:200] if len(name) > 200 else name


def save_upload_to_disk(file_bytes: bytes, original_name: str) -> Tuple[str, str]:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    fn = safe_filename(original_name)
    uid = secrets.token_hex(16)
    stored = f"{uid}__{fn}"
    path = os.path.join(UPLOAD_DIR, stored)
    with open(path, "wb") as f:
        f.write(file_bytes)
    return stored, path


@app.post("/requests/{request_id}/attachments/upload")
async def request_attachment_upload(
    request: Request,
    request_id: int,
    file: UploadFile = File(...),
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)),
    db: Session = Depends(db_session),
):
    req = db.get(PurchaseRequest, request_id)
    if not req:
        raise HTTPException(404, "Request not found")

    data = await file.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(400, f"Файл слишком большой (>{MAX_UPLOAD_MB}MB)")

    stored_name, _ = save_upload_to_disk(data, file.filename or "file")
    att = Attachment(
        kind="REQUEST",
        request_id=req.id,
        original_name=file.filename or "file",
        stored_path=stored_name,
        content_type=file.content_type or "",
        size_bytes=len(data),
    )
    db.add(att)
    db.commit()
    audit(db, request, "REQUEST_ATTACHMENT_UPLOAD", "Request", req.id, att.original_name, user=user)
    return RedirectResponse(f"/requests/{request_id}?tab=docs", status_code=303)


# -----------------------------
# Rounds + invitations
# -----------------------------
@app.post("/requests/{request_id}/rounds/create")
def round_create(
    request: Request,
    request_id: int,
    deadline_at: str = Form(...),  # "YYYY-MM-DDTHH:MM"
    comment_to_suppliers: str = Form(""),
    type_value: str = Form(RoundType.INITIAL),
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)),
    db: Session = Depends(db_session),
):
    req = db.get(PurchaseRequest, request_id)
    if not req:
        raise HTTPException(404, "Request not found")

    # round number = max+1
    max_no = db.scalar(select(func.max(Round.number)).where(Round.request_id == request_id)) or 0
    rn = int(max_no) + 1

    # parse local naive, treat as UTC for simplicity (для MVP)
    dt = datetime.fromisoformat(deadline_at)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    if dt <= now_utc():
        raise HTTPException(400, "Дедлайн должен быть в будущем")

    rnd = Round(
        request_id=request_id,
        number=rn,
        type=type_value,
        deadline_at=dt,
        comment_to_suppliers=comment_to_suppliers.strip(),
        created_by_user_id=user.id,
        is_closed=False
    )
    db.add(rnd)
    db.commit()
    db.refresh(rnd)
    # participants default = category suppliers
    cat = db.get(Category, req.category_id)
    default_suppliers = [s for s in (cat.suppliers if cat else []) if s.status == SupplierStatus.ACTIVE]

    for s in default_suppliers:
        exists = db.scalar(select(RoundSupplier).where(RoundSupplier.round_id == rnd.id, RoundSupplier.supplier_id == s.id))
        if not exists:
            db.add(RoundSupplier(round_id=rnd.id, supplier_id=s.id))
    db.commit()

    # Обновление статуса заявки
    if rn == 1:
        req.status = RequestStatus.ROUND_1_OPEN
    else:
        req.status = RequestStatus.ROUND_N_OPEN
    db.commit()

    audit(db, request, "ROUND_CREATE", "Request", req.id, f"round={rn}", user=user)
    return RedirectResponse(f"/requests/{request_id}?tab=rounds", status_code=303)


@app.post("/rounds/{round_id}/close")
def round_close(
    request: Request,
    round_id: int,
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)),
    db: Session = Depends(db_session),
):
    rnd = db.get(Round, round_id)
    if not rnd:
        raise HTTPException(404, "Round not found")
    rnd.is_closed = True
    rnd.closed_at = now_utc()
    db.commit()

    req = db.get(PurchaseRequest, rnd.request_id)
    if rnd.number == 1:
        req.status = RequestStatus.ROUND_1_CLOSED
    else:
        req.status = RequestStatus.ROUND_N_CLOSED
    db.commit()

    audit(db, request, "ROUND_CLOSE", "Request", req.id, f"round={rnd.number}", user=user)
    return RedirectResponse(f"/requests/{rnd.request_id}?tab=rounds", status_code=303)


@app.post("/rounds/{round_id}/send_invites")
def round_send_invites(
    request: Request,
    round_id: int,
    background: BackgroundTasks,
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)),
    db: Session = Depends(db_session),
):
    rnd = db.get(Round, round_id)
    if not rnd:
        raise HTTPException(404, "Round not found")
    req = db.get(PurchaseRequest, rnd.request_id)
    cat = db.get(Category, req.category_id)

    # suppliers by category (FR-10)
    participants = db.scalars(select(RoundSupplier).where(RoundSupplier.round_id == rnd.id)).all()
    supplier_ids = [p.supplier_id for p in participants]
    suppliers = db.scalars(select(Supplier).where(Supplier.id.in_(supplier_ids))).all()
    suppliers = [s for s in suppliers if s.status == SupplierStatus.ACTIVE]

    if not suppliers:
        raise HTTPException(400, "В раунде нет активных участников (задай участников раунда)")

    sent_count = 0
    for s in suppliers:
        # ensure invitation exists
        inv = db.scalar(select(Invitation).where(Invitation.round_id == rnd.id, Invitation.supplier_id == s.id))
        if not inv:
            token = secrets.token_urlsafe(32)
            inv = Invitation(
                round_id=rnd.id,
                supplier_id=s.id,
                token_hash=token_hash(token),
                status=InvitationStatus.CREATED
            )
            db.add(inv)
            db.commit()
            db.refresh(inv)
        else:
            # Если уже было — не генерим новый токен (чтобы не путать поставщика)
            token = None

        # отправка на все email поставщика
        emails = [e.strip() for e in (s.emails or "").split(",") if e.strip()]
        if not emails:
            continue

        if token is None:
            # токен уже был, но у нас его нет (мы не храним). Для простоты:
            # создадим новый токен и заменим хеш (чтобы можно было снова отправить).
            token = secrets.token_urlsafe(32)
            inv.token_hash = token_hash(token)
            inv.status = InvitationStatus.CREATED
            db.commit()

        link = make_public_link(token)
        subject, text = build_invite_email(req, rnd, link)

        for e in emails:
            background.add_task(send_email_smtp, e, subject, text)

        inv.status = InvitationStatus.SENT
        inv.sent_at = now_utc()
        db.commit()
        sent_count += 1

    audit(db, request, "ROUND_SEND_INVITES", "Request", req.id, f"round={rnd.number}, sent={sent_count}", user=user)
    return RedirectResponse(f"/requests/{req.id}?tab=rounds", status_code=303)


@app.post("/rounds/{round_id}/participants/set")
def round_participants_set(
    request: Request,
    round_id: int,
    supplier_ids: str = Form(""),
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)),
    db: Session = Depends(db_session),
):
    rnd = db.get(Round, round_id)
    if not rnd:
        raise HTTPException(404, "Round not found")

    ids = [int(x) for x in re.findall(r"\d+", supplier_ids or "")]
    # очистить старых
    db.query(RoundSupplier).where(RoundSupplier.round_id == rnd.id).delete()
    db.commit()

    if ids:
        valid = db.scalars(select(Supplier).where(Supplier.id.in_(ids))).all()
        for s in valid:
            if s.status == SupplierStatus.ACTIVE:
                db.add(RoundSupplier(round_id=rnd.id, supplier_id=s.id))
        db.commit()

    audit(db, request, "ROUND_PARTICIPANTS_SET", "Round", rnd.id, f"ids={ids}", user=user)
    return RedirectResponse(f"/requests/{rnd.request_id}?tab=rounds", status_code=303)


# -----------------------------
# Public supplier form (token link)
# -----------------------------
def find_invitation_by_token(db: Session, token: str) -> Invitation:
    th = token_hash(token)
    inv = db.scalar(select(Invitation).where(Invitation.token_hash == th))
    if not inv:
        raise HTTPException(404, "Ссылка недействительна")
    return inv


def get_previous_round_offer(db: Session, req_id: int, current_round_number: int, supplier_id: int) -> Optional[Offer]:
    if current_round_number <= 1:
        return None
    prev_round = db.scalar(select(Round).where(Round.request_id == req_id, Round.number == (current_round_number - 1)))
    if not prev_round:
        return None
    return get_round_offer(db, prev_round.id, supplier_id)


@app.get("/public/offer/{token}", response_class=HTMLResponse)
def public_offer_form(
    request: Request,
    token: str,
    db: Session = Depends(db_session),
):
    rate_limit_public(request)

    inv = find_invitation_by_token(db, token)
    rnd = db.get(Round, inv.round_id)
    req = db.get(PurchaseRequest, rnd.request_id)
    cat = db.get(Category, req.category_id)
    supplier = db.get(Supplier, inv.supplier_id)

    # ✅ ВСЕГДА определяем вложения заявки
    request_attachments = db.scalars(
        select(Attachment)
        .where(Attachment.kind == "REQUEST", Attachment.request_id == req.id)
        .order_by(Attachment.created_at.desc())
    ).all()

    close_round_if_deadline_passed(db, rnd)
    if not round_is_open(rnd):
        inv.status = InvitationStatus.EXPIRED
        db.commit()
        return templates.TemplateResponse("supplier_offer.html", {
            "request": request,
            "token": token,
            "mode": "closed",
            "supplier": supplier,
            "req": req,
            "rnd": rnd,
            "cat": cat,
            "request_attachments": request_attachments,
        }, status_code=400)

    if inv.opened_at is None:
        inv.opened_at = now_utc()
        if inv.status in (InvitationStatus.SENT, InvitationStatus.CREATED):
            inv.status = InvitationStatus.OPENED
        db.commit()

    items = db.scalars(
        select(RequestItem).where(RequestItem.request_id == req.id).order_by(RequestItem.pos_no.asc())
    ).all()
    existing_offer = get_round_offer(db, rnd.id, supplier.id)

    prev_offer = get_previous_round_offer(db, req.id, rnd.number, supplier.id)
    prev_map = {}
    if prev_offer:
        prev_items = db.scalars(select(OfferItem).where(OfferItem.offer_id == prev_offer.id)).all()
        prev_map = {x.request_item_id: x for x in prev_items}

    cur_map = {}
    if existing_offer:
        cur_items = db.scalars(select(OfferItem).where(OfferItem.offer_id == existing_offer.id)).all()
        cur_map = {x.request_item_id: x for x in cur_items}

    return templates.TemplateResponse("supplier_offer.html", {
        "request": request,
        "token": token,
        "mode": "open",
        "supplier": supplier,
        "req": req,
        "rnd": rnd,
        "cat": cat,
        "items": items,
        "existing_offer": existing_offer,
        "cur_map": cur_map,
        "prev_offer": prev_offer,
        "prev_map": prev_map,
        "require_all_positions": True,
        "request_attachments": request_attachments,
    })


@app.post("/public/offer/{token}")
async def public_offer_submit(
    request: Request,
    token: str,
    payment_terms: str = Form(""),
    valid_until: str = Form(""),
    comment: str = Form(""),
    # вложения поставщика (опционально)
    attachment: Optional[UploadFile] = File(None),
    db: Session = Depends(db_session),
):
    rate_limit_public(request)

    inv = find_invitation_by_token(db, token)
    rnd = db.get(Round, inv.round_id)
    req = db.get(PurchaseRequest, rnd.request_id)
    cat = db.get(Category, req.category_id)
    supplier = db.get(Supplier, inv.supplier_id)

    close_round_if_deadline_passed(db, rnd)
    if not round_is_open(rnd):
        inv.status = InvitationStatus.EXPIRED
        db.commit()
        raise HTTPException(400, "Раунд закрыт или истёк дедлайн (отправка запрещена)")

    if cat and cat.require_payment_terms and not payment_terms.strip():
        raise HTTPException(400, "Условия оплаты обязательны для этой категории")

    items = db.scalars(select(RequestItem).where(RequestItem.request_id == req.id).order_by(RequestItem.pos_no.asc())).all()
    if not items:
        raise HTTPException(400, "В заявке нет позиций")

    # offer create/update
    offer = get_round_offer(db, rnd.id, supplier.id)
    if not offer:
        offer = Offer(round_id=rnd.id, supplier_id=supplier.id)
        db.add(offer)
        db.commit()
        db.refresh(offer)

    offer.payment_terms = payment_terms.strip()
    offer.valid_until = valid_until.strip()
    offer.comment = comment.strip()

    # Для простоты: принимаем поля вида:
    # price_<request_item_id>, days_<id>, ns_<id>
    forbid_inc = bool(cat.forbid_price_increase) if cat else False
    allow_ns = bool(cat.allow_not_supply) if cat else True

    # previous offer (для запрета повышения цены в переторжке)
    prev_offer = get_previous_round_offer(db, req.id, rnd.number, supplier.id)
    prev_map = {}
    if prev_offer:
        prev_items = db.scalars(select(OfferItem).where(OfferItem.offer_id == prev_offer.id)).all()
        prev_map = {x.request_item_id: x for x in prev_items}

    # очищаем старые items и создаём заново — проще для новичка
    db.query(OfferItem).where(OfferItem.offer_id == offer.id).delete()
    db.commit()

    missing = 0
    for ri in items:
        key_price = f"price_{ri.id}"
        key_days = f"days_{ri.id}"
        key_ns = f"ns_{ri.id}"

        form = await request.form()
        price_val = form.get(key_price)
        days_val = form.get(key_days)
        ns_val = form.get(key_ns)

        if price_val is None or str(price_val).strip() == "":
            missing += 1
            continue

        unit_price = float(str(price_val).replace(",", "."))
        delivery_days = None
        if days_val and str(days_val).strip():
            delivery_days = int(str(days_val).strip())

        not_supply = False
        if allow_ns and ns_val in ("on", "true", "1", "yes"):
            not_supply = True

        # FR-22: запрет повышения цены в переторжке (если включено)
        if forbid_inc and prev_map.get(ri.id) and not not_supply:
            prev_price = float(prev_map[ri.id].unit_price)
            if unit_price > prev_price + 1e-9:
                raise HTTPException(400, f"Повышение цены запрещено (позиция {ri.pos_no}). Было: {prev_price}, стало: {unit_price}")

        oi = OfferItem(
            offer_id=offer.id,
            request_item_id=ri.id,
            unit_price=unit_price,
            delivery_days=delivery_days,
            comment=str(form.get(f"c_{ri.id}") or "").strip(),
            not_supply=not_supply
        )
        db.add(oi)

    if missing > 0:
        raise HTTPException(400, f"Нужно заполнить цену по всем позициям. Не заполнено: {missing}")

    db.commit()
    db.refresh(offer)
    compute_offer_totals(db, offer)
    db.commit()

    # attachment (optional)
    if attachment is not None:
        data = await attachment.read()
        if len(data) > MAX_UPLOAD_BYTES:
            raise HTTPException(400, f"Вложение слишком большое (>{MAX_UPLOAD_MB}MB)")
        stored_name, _ = save_upload_to_disk(data, attachment.filename or "file")
        att = Attachment(
            kind="OFFER",
            offer_id=offer.id,
            original_name=attachment.filename or "file",
            stored_path=stored_name,
            content_type=attachment.content_type or "",
            size_bytes=len(data),
        )
        db.add(att)
        db.commit()

    inv.status = InvitationStatus.RESPONDED
    inv.responded_at = now_utc()
    db.commit()

    audit(db, request, "SUPPLIER_OFFER_SUBMIT", "Request", req.id, f"round={rnd.number}, supplier={supplier.name}", supplier=supplier)
    return templates.TemplateResponse("supplier_offer.html", {
        "request": request,
        "token": token,
        "mode": "done",
        "supplier": supplier,
        "req": req,
        "rnd": rnd,
        "cat": cat
    })


# -----------------------------
# Compare table (internal)
# -----------------------------
def build_compare_table(db: Session, req: PurchaseRequest, round_id: int) -> Dict:
    items = db.scalars(select(RequestItem).where(RequestItem.request_id == req.id).order_by(RequestItem.pos_no.asc())).all()
    offers = db.scalars(select(Offer).where(Offer.round_id == round_id)).all()
    suppliers = {o.supplier_id: db.get(Supplier, o.supplier_id) for o in offers}

    offer_items = {}
    for o in offers:
        ois = db.scalars(select(OfferItem).where(OfferItem.offer_id == o.id)).all()
        offer_items[o.id] = {x.request_item_id: x for x in ois}

    # best per line (min unit_price among not_supply==False)
    best_unit_price: Dict[int, float] = {}
    for ri in items:
        prices = []
        for o in offers:
            oi = offer_items[o.id].get(ri.id)
            if not oi or oi.not_supply:
                continue
            prices.append(float(oi.unit_price))
        best_unit_price[ri.id] = min(prices) if prices else None

    rows = []
    for o in offers:
        rows.append({
            "offer": o,
            "supplier": suppliers.get(o.supplier_id),
            "items": offer_items.get(o.id, {})
        })

    return {
        "items": items,
        "rows": rows,
        "best_unit_price": best_unit_price
    }


# -----------------------------
# Reports: PDF + XLSX snapshots
# -----------------------------
def _fig_to_png_bytes(fig) -> bytes:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=160, bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


def make_charts(db: Session, req: PurchaseRequest, rnd: Round) -> Dict[str, bytes]:
    offers = db.scalars(select(Offer).where(Offer.round_id == rnd.id).order_by(Offer.total_amount.asc().nulls_last())).all()
    labels = []
    totals = []
    scatter_x = []
    scatter_y = []

    for o in offers:
        s = db.get(Supplier, o.supplier_id)
        labels.append(s.name if s else f"Supplier {o.supplier_id}")
        totals.append(float(o.total_amount or 0))
        # scatter: x=total price, y=avg delivery days (optional)
        ois = db.scalars(select(OfferItem).where(OfferItem.offer_id == o.id)).all()
        days = [oi.delivery_days for oi in ois if oi.delivery_days is not None and not oi.not_supply]
        avg_days = sum(days) / len(days) if days else 0
        scatter_x.append(float(o.total_amount or 0))
        scatter_y.append(avg_days)

    # Bar totals
    fig1 = plt.figure(figsize=(8, 3.5))
    ax = fig1.add_subplot(111)
    ax.bar(range(len(labels)), totals, color="#2563eb")
    ax.set_title("Итоговые суммы по поставщикам")
    ax.set_ylabel(f"Сумма ({req.currency})")
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=30, ha="right")
    chart_totals = _fig_to_png_bytes(fig1)

    # Scatter
    fig2 = plt.figure(figsize=(8, 3.5))
    ax2 = fig2.add_subplot(111)
    ax2.scatter(scatter_x, scatter_y, color="#16a34a")
    ax2.set_title("Цена vs срок (scatter)")
    ax2.set_xlabel(f"Итог ({req.currency})")
    ax2.set_ylabel("Средний срок (дней)")
    chart_scatter = _fig_to_png_bytes(fig2)

    # Round1 vs Round2 (если есть)
    chart_r1r2 = None
    if rnd.number >= 2:
        r1 = db.scalar(select(Round).where(Round.request_id == req.id, Round.number == 1))
        if r1:
            offers1 = db.scalars(select(Offer).where(Offer.round_id == r1.id)).all()
            map1 = {o.supplier_id: float(o.total_amount or 0) for o in offers1}
            map2 = {o.supplier_id: float(o.total_amount or 0) for o in offers}
            common_ids = sorted(set(map1.keys()) | set(map2.keys()))
            labs = []
            v1 = []
            v2 = []
            for sid in common_ids:
                s = db.get(Supplier, sid)
                labs.append(s.name if s else str(sid))
                v1.append(map1.get(sid, 0))
                v2.append(map2.get(sid, 0))
            fig3 = plt.figure(figsize=(8, 3.5))
            ax3 = fig3.add_subplot(111)
            x = range(len(labs))
            ax3.bar([i - 0.2 for i in x], v1, width=0.4, label="Раунд 1", color="#64748b")
            ax3.bar([i + 0.2 for i in x], v2, width=0.4, label=f"Раунд {rnd.number}", color="#f97316")
            ax3.set_title("Сравнение Раунд 1 vs финальный")
            ax3.set_ylabel(f"Сумма ({req.currency})")
            ax3.set_xticks(list(x))
            ax3.set_xticklabels(labs, rotation=30, ha="right")
            ax3.legend()
            chart_r1r2 = _fig_to_png_bytes(fig3)

    return {
        "totals": chart_totals,
        "scatter": chart_scatter,
        "r1r2": chart_r1r2
    }


def generate_pdf_bytes(db: Session, req: PurchaseRequest, rnd: Round) -> Tuple[bytes, Dict]:
    styles = getSampleStyleSheet()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=15*mm, bottomMargin=15*mm)

    items = db.scalars(select(RequestItem).where(RequestItem.request_id == req.id).order_by(RequestItem.pos_no.asc())).all()
    offers = db.scalars(select(Offer).where(Offer.round_id == rnd.id).order_by(Offer.total_amount.asc().nulls_last())).all()

    # Snapshot payload (простая версия FR-30)
    snapshot = {
        "request": {"id": req.id, "number": req.number, "subject": req.subject, "currency": req.currency, "vat_mode": req.vat_mode},
        "round": {"id": rnd.id, "number": rnd.number, "deadline_at": rnd.deadline_at.isoformat()},
        "generated_at": now_utc().isoformat(),
        "offers": []
    }

    story = []
    story.append(Paragraph(f"<b>Аналитическая справка</b>", styles["Title"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Заявка: <b>{req.number}</b> — {req.subject}", styles["Normal"]))
    story.append(Paragraph(f"Раунд: <b>{rnd.number}</b> ({rnd.type})", styles["Normal"]))
    story.append(Paragraph(f"Дедлайн: {rnd.deadline_at.strftime('%Y-%m-%d %H:%M %Z')}", styles["Normal"]))
    story.append(Spacer(1, 10))

    # Summary table
    data = [["Поставщик", f"Итог ({req.currency})", "Комментарий"]]
    for o in offers:
        s = db.get(Supplier, o.supplier_id)
        data.append([s.name if s else str(o.supplier_id), f"{float(o.total_amount or 0):,.2f}", (o.comment or "")[:80]])
        snapshot["offers"].append({
            "supplier_id": o.supplier_id,
            "supplier_name": s.name if s else None,
            "total": float(o.total_amount or 0)
        })

    tbl = Table(data, colWidths=[75*mm, 35*mm, 65*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP")
    ]))
    story.append(Paragraph("<b>Сводная таблица финального раунда</b>", styles["Heading2"]))
    story.append(tbl)
    story.append(Spacer(1, 12))

    # Charts
    charts = make_charts(db, req, rnd)
    story.append(Paragraph("<b>Графики</b>", styles["Heading2"]))

    img1 = Image(io.BytesIO(charts["totals"]), width=170*mm, height=70*mm)
    story.append(img1)
    story.append(Spacer(1, 8))

    img2 = Image(io.BytesIO(charts["scatter"]), width=170*mm, height=70*mm)
    story.append(img2)
    story.append(Spacer(1, 8))

    if charts["r1r2"]:
        img3 = Image(io.BytesIO(charts["r1r2"]), width=170*mm, height=70*mm)
        story.append(img3)
        story.append(Spacer(1, 8))

    story.append(PageBreak())

    # Items detail (first supplier only as example? no — даём просто перечень позиций)
    story.append(Paragraph("<b>Позиции заявки</b>", styles["Heading2"]))
    item_data = [["№", "Наименование", "Кол-во", "Ед.изм."]]
    for ri in items[:200]:  # ограничим, чтобы PDF не раздувался бесконечно
        item_data.append([str(ri.pos_no), ri.name[:60], str(float(ri.qty)), ri.uom])
    item_tbl = Table(item_data, colWidths=[10*mm, 120*mm, 20*mm, 20*mm])
    item_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    story.append(item_tbl)

    doc.build(story)
    return buf.getvalue(), snapshot


def generate_xlsx_bytes(db: Session, req: PurchaseRequest) -> Tuple[bytes, Dict]:
    wb = Workbook()
    wb.remove(wb.active)

    snapshot = {"generated_at": now_utc().isoformat(), "request_id": req.id, "rounds": []}

    # Request sheet
    ws = wb.create_sheet("Request")
    ws.append(["Number", req.number])
    ws.append(["Subject", req.subject])
    ws.append(["Currency", req.currency])
    ws.append(["VAT Mode", req.vat_mode])
    ws.append(["Payment terms", req.payment_terms])
    ws.append(["Delivery terms", req.delivery_terms])
    ws.append(["Status", req.status])
    ws.append(["Created at", req.created_at.isoformat()])

    # Items
    ws = wb.create_sheet("Items")
    ws.append(["pos_no", "name", "description", "qty", "uom"])
    items = db.scalars(select(RequestItem).where(RequestItem.request_id == req.id).order_by(RequestItem.pos_no.asc())).all()
    for ri in items:
        ws.append([ri.pos_no, ri.name, ri.description, float(ri.qty), ri.uom])

    # Rounds + offers
    rounds = db.scalars(select(Round).where(Round.request_id == req.id).order_by(Round.number.asc())).all()
    for rnd in rounds:
        offers = db.scalars(select(Offer).where(Offer.round_id == rnd.id)).all()
        ws = wb.create_sheet(f"Round_{rnd.number}_Offers")
        ws.append(["supplier", "item_pos", "item_name", "unit_price", "qty", "line_amount", "delivery_days", "not_supply"])
        for o in offers:
            s = db.get(Supplier, o.supplier_id)
            ois = db.scalars(select(OfferItem).where(OfferItem.offer_id == o.id)).all()
            oi_map = {x.request_item_id: x for x in ois}
            for ri in items:
                oi = oi_map.get(ri.id)
                if not oi:
                    continue
                qty = float(ri.qty)
                unit_price = float(oi.unit_price)
                line_amount = 0.0 if oi.not_supply else qty * unit_price
                ws.append([
                    s.name if s else str(o.supplier_id),
                    ri.pos_no,
                    ri.name,
                    unit_price,
                    qty,
                    round(line_amount, 2),
                    oi.delivery_days,
                    bool(oi.not_supply)
                ])

        snapshot["rounds"].append({"round": rnd.number, "offers": len(offers)})

    # Summary
    ws = wb.create_sheet("Summary")
    ws.append(["Round", "Supplier", "Total"])
    for rnd in rounds:
        offers = db.scalars(select(Offer).where(Offer.round_id == rnd.id).order_by(Offer.total_amount.asc().nulls_last())).all()
        for o in offers:
            s = db.get(Supplier, o.supplier_id)
            ws.append([rnd.number, s.name if s else str(o.supplier_id), float(o.total_amount or 0)])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), snapshot


@app.post("/requests/{request_id}/reports/generate")
def reports_generate(
    request: Request,
    request_id: int,
    round_id: int = Form(...),
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER, Role.VIEWER)),
    db: Session = Depends(db_session),
):
    req = db.get(PurchaseRequest, request_id)
    rnd = db.get(Round, round_id)
    if not req or not rnd or rnd.request_id != req.id:
        raise HTTPException(404, "Not found")

    pdf_bytes, pdf_snapshot = generate_pdf_bytes(db, req, rnd)
    xlsx_bytes, xlsx_snapshot = generate_xlsx_bytes(db, req)

    pdf = Report(
        request_id=req.id,
        round_id=rnd.id,
        kind="PDF",
        file_bytes=pdf_bytes,
        filename=f"{req.number}_round{rnd.number}_report.pdf",
        snapshot_json=json.dumps(pdf_snapshot, ensure_ascii=False),
        created_by_user_id=user.id
    )
    xlsx = Report(
        request_id=req.id,
        round_id=rnd.id,
        kind="XLSX",
        file_bytes=xlsx_bytes,
        filename=f"{req.number}_export.xlsx",
        snapshot_json=json.dumps(xlsx_snapshot, ensure_ascii=False),
        created_by_user_id=user.id
    )
    db.add_all([pdf, xlsx])
    db.commit()

    audit(db, request, "REPORT_GENERATE", "Request", req.id, f"round={rnd.number}", user=user)
    return RedirectResponse(f"/requests/{request_id}?tab=reports&round_id={round_id}", status_code=303)


@app.get("/reports/{report_id}/download")
def report_download(
    request: Request,
    report_id: int,
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER, Role.VIEWER)),
    db: Session = Depends(db_session),
):
    rep = db.get(Report, report_id)
    if not rep:
        raise HTTPException(404, "Report not found")

    media = "application/pdf" if rep.kind == "PDF" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    headers = {"Content-Disposition": f'attachment; filename="{rep.filename}"'}
    return Response(content=rep.file_bytes, media_type=media, headers=headers)


# -----------------------------
# Decision: choose winner (simple)
# -----------------------------
@app.post("/requests/{request_id}/decision")
def set_decision(
    request: Request,
    request_id: int,
    round_id: int = Form(...),
    winner_supplier_id: int = Form(...),
    reason: str = Form(""),
    user: User = Depends(require_roles(Role.ADMIN, Role.BUYER)),
    db: Session = Depends(db_session),
):
    req = db.get(PurchaseRequest, request_id)
    rnd = db.get(Round, round_id)
    if not req or not rnd or rnd.request_id != req.id:
        raise HTTPException(404, "Not found")

    req.status = RequestStatus.DECISION
    db.commit()

    audit(db, request, "DECISION_SET", "Request", req.id, f"round={rnd.number}, winner_supplier_id={winner_supplier_id}, reason={reason}", user=user)
    return RedirectResponse(f"/requests/{request_id}?tab=offers&round_id={round_id}", status_code=303)


# -----------------------------
# CLI: init admin
# -----------------------------
def init_admin():
    Base.metadata.create_all(engine)
    admin_email = os.getenv("ADMIN_EMAIL", "admin@example.com").lower().strip()
    admin_password = os.getenv("ADMIN_PASSWORD", "admin12345")
    admin_name = os.getenv("ADMIN_NAME", "Admin")

    with Session(engine) as db:
        u = db.scalar(select(User).where(User.email == admin_email))
        if u:
            print("Admin already exists:", admin_email)
            return
        u = User(
            email=admin_email,
            name=admin_name,
            password_hash=hash_password(admin_password),
            role=Role.ADMIN,
            is_active=True
        )
        db.add(u)
        db.commit()
        print("Admin created:", admin_email)
        print("Password:", admin_password)


def create_user_cli(email: str, password: str, role: str, name: str = ""):
    Base.metadata.create_all(engine)

    email = (email or "").strip().lower()
    name = (name or "").strip()
    role = (role or "").strip().upper()

    if role not in (Role.ADMIN, Role.BUYER, Role.VIEWER):
        raise SystemExit(f"Invalid role: {role}. Use ADMIN, BUYER, VIEWER")

    if len(password.encode("utf-8")) > 72:
        raise SystemExit("Password too long for bcrypt (must be <= 72 bytes). Use shorter password.")

    with Session(engine) as db:
        existing = db.scalar(select(User).where(User.email == email))
        if existing:
            raise SystemExit(f"User already exists: {email}")

        u = User(
            email=email,
            name=name,
            password_hash=hash_password(password),
            role=role,
            is_active=True,
        )
        db.add(u)
        db.commit()
        print("User created:", email, "role=", role)


def send_reminders_cli(hours_before: int = 24, cooldown_hours: int = 12):
    Base.metadata.create_all(engine)
    cutoff = now_utc() + timedelta(hours=hours_before)
    cooldown = now_utc() - timedelta(hours=cooldown_hours)

    with Session(engine) as db:
        # open rounds, deadline soon
        rounds = db.scalars(
            select(Round)
            .where(Round.is_closed == False, Round.deadline_at <= cutoff, Round.deadline_at >= now_utc())
        ).all()

        sent = 0
        for rnd in rounds:
            req = db.get(PurchaseRequest, rnd.request_id)

            invs = db.scalars(
                select(Invitation).where(
                    Invitation.round_id == rnd.id,
                    Invitation.status.in_([InvitationStatus.SENT, InvitationStatus.OPENED]),
                    (Invitation.reminded_at.is_(None)) | (Invitation.reminded_at < cooldown),
                )
            ).all()

            for inv in invs:
                supplier = db.get(Supplier, inv.supplier_id)
                emails = [e.strip() for e in (supplier.emails or "").split(",") if e.strip()]
                if not emails:
                    continue

                # генерим новый токен (мы не храним исходный), обновляем hash
                token = secrets.token_urlsafe(32)
                inv.token_hash = token_hash(token)
                link = make_public_link(token)

                subj, body = build_reminder_email(req, rnd, link)
                for e in emails:
                    send_email_smtp(e, subj, body)

                # записываем reminded_at (через raw SQL, раз колонка добавлена ALTER-ом)
                inv.reminded_at = now_utc()
                db.commit()
                sent += 1

        print("Reminders sent:", sent)



if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--init-admin", action="store_true")

    parser.add_argument("--create-user", action="store_true")
    parser.add_argument("--email", type=str, default="")
    parser.add_argument("--password", type=str, default="")
    parser.add_argument("--role", type=str, default=Role.VIEWER)
    parser.add_argument("--name", type=str, default="")
    parser.add_argument("--send-reminders", action="store_true")
    parser.add_argument("--reminder-hours", type=int, default=24)

    args = parser.parse_args()

    if args.init_admin:
        init_admin()
    elif args.create_user:
        if not args.email or not args.password:
            raise SystemExit("Usage: --create-user --email ... --password ... [--role ADMIN|BUYER|VIEWER] [--name ...]")
        create_user_cli(email=args.email, password=args.password, role=args.role, name=args.name)
    elif args.send_reminders:
        send_reminders_cli(hours_before=args.reminder_hours)
    else:
        print("Run with: uvicorn app:app --reload")

